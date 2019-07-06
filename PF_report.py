import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
    import pandas as pd
    import numpy as np
    import xlrd
except:
    install('pandas')
    install('numpy')
    install('xlrd')

import pandas as pd
import numpy as np
import xlrd

# In[78]:

excel_df_name = str(input("Please give the name of the file with report. The data should be on the first tab.\n"))
excel_df = pd.read_excel(excel_df_name)
excel_df


# In[79]:


excel_df["Сред. длительность сеанса"] = excel_df["Сред. длительность сеанса"]/86400
excel_df["Новые посещения %"] = excel_df["Новые пользователи"]/excel_df["Пользователи"]


# In[80]:


wm_by_session = lambda x: np.average(x, weights=excel_df.loc[x.index, "Сеансы"])

by_channel = excel_df.groupby('Источник или канал').agg({
    "Сеансы": sum,
    "Сред. длительность сеанса": wm_by_session,
    "Страниц/сеанс": wm_by_session,
    "Новые посещения %": wm_by_session,
    "Показатель отказов": wm_by_session,
    "Транзакции":sum,
    "Доход": sum
    })                                                  


# In[81]:


by_channel_and_campaign = excel_df.groupby(['Источник или канал', "Кампания"]).agg({
    "Сеансы": sum,
    "Сред. длительность сеанса": wm_by_session,
    "Страниц/сеанс": wm_by_session,
    "Новые посещения %": wm_by_session,
    "Показатель отказов": wm_by_session,
    "Транзакции":sum,
    "Доход": sum
    })  


# In[82]:


by_channel.to_excel("grouped_by_channel.xlsx", merge_cells=False )
by_channel_and_campaign.to_excel("grouped_by_channel_and_campaign.xlsx", merge_cells=False)


# In[83]:


"""
#%load file.py

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

"""

